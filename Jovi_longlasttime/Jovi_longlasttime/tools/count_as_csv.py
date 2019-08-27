import logging
import os
import shutil
import time
import zipfile

from openpyxl import load_workbook

date = time.strftime('%m-%d', time.localtime())
path = 'e:\数据统计.xlsx'
dir = 'e:\定期更新{}'.format(date)
zip_file = 'e:\\定期更新%s.zip' % date

os.chdir('E:\\')
book = load_workbook(filename=path)

web = ['IT之家', 'Zaker新闻', '东方头条', '今日头条', '凤凰网', '手机网易网', '手机新浪网', '搜狗微信', '搜狐网', '腾讯新闻',
       '微博长文', '微博短文', '新华网', '央视新闻', '一点资讯', '汽车之家', '人民网']  # 共15个网站


# 对单个文件夹进行遍历统计新闻数量
def all_count(path):
    nums = 0

    def count(_path):
        nonlocal nums
        if os.path.isdir(_path):
            for i in os.listdir(_path):
                os.chdir(_path)
                sub_path = os.path.join(_path, i)
                count(sub_path)
        else:
            with open(_path, 'r', encoding='utf-8') as f:
                articles = len(f.readlines())
                nums += articles
        return nums

    os.chdir(path)
    all_nums = dict()
    zs = 0
    for i in os.listdir(path):
        _i = i.rstrip('.txt')
        complete_path = os.path.join(path, i)
        all_nums[_i] = count(complete_path)
        zs += all_nums[_i]
        nums = 0
    all_nums['总数'] = zs
    return all_nums


# 新建文件夹，并将各个网站新闻移动到这个文件夹
def move_dir(dir):
    if not os.path.exists(dir):
        os.mkdir(dir)
    os.chdir(dir)
    for i in web:
        src = 'e:\\{}'.format(i)
        dst = dir
        try:
            shutil.move(src, dst)
            logging.info('move {} to {} DONE'.format(src, dst))
        except Exception as e:
            logging.warning('src not exist:{}'.format(src))
            print(e)
            continue
    logging.info('move all srcs!')


def remove_csv(dir, date):
    src = path
    dst = '{}\\数据统计{}.xlsx'.format(dir, date)
    try:
        shutil.copy2(src, dst)
    except Exception as e:
        logging.error('excel文件复制错误')
        print(e)


def count_web(path):
    return all_count(path)


def read_and_write(sheet_name, article_nums):
    if sheet_name not in book.sheetnames:
        sheet = book.create_sheet(sheet_name)
    else:
        sheet = book.get_sheet_by_name(sheet_name)
    keys = list(article_nums.keys())
    keys.reverse()
    values = list(article_nums.values())
    values.reverse()
    sheet.append(['日期'] + keys)
    sheet.append([date] + values)
    return


def main():
    for subdir in os.listdir(dir):
        new_path = os.path.join(dir, subdir)
        count = count_web(new_path)
        read_and_write(subdir, count)
    book.save(path)


def file_generator(input_dir):
    for i in os.walk(input_dir):
        a, d, f = i
        for _d in d:
            file_generator(_d)
        for _f in f:
            yield os.path.join(a, _f)


def zip(src, dst):
    f = zipfile.ZipFile(dst, 'w', zipfile.ZIP_DEFLATED)
    file_list = file_generator(src)
    for file in file_list:
        f.write(file)
    f.close()


if __name__ == '__main__':
    move_dir(dir)
    main()
    remove_csv(dir, date)
    zip(dir, zip_file)
